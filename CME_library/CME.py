from openpyxl import load_workbook
import struct
import socket

def int_number_to_int_list(string_number, byte_length):
    """
    Convert an integer to a hexadecimal string in little-endian format with specified byte length.
    """
    packed_bytes = struct.pack(f'<{byte_length}s', struct.pack('<i', string_number)[:byte_length])
    return [byte for byte in packed_bytes]

def calculate_crc_add_magicNumber(int_data_list):
    # 将十六进制数求和
    sum_value = sum(int_data_list)
    #sum_value = hex(sum_value)
    # 按位取反操作
    crc_result_0 = sum_value ^ 0xFF
    # 取后8位
    crc_result = crc_result_0 & 0xFF
    int_data_list.append(crc_result)
    int_data_list.insert(0, 90)

    return int_data_list

def generate_button_control_request_command(excel_file, moudle_id):
    button_control_command_dict = {}
    # 加载Excel文件
    workbook = load_workbook(filename=excel_file, data_only=True)
    # 选择工作表
    button_control_sheet = workbook['Button Control']

    Button_number = button_control_sheet.cell(row=4, column=7).value or button_control_sheet.cell(row=4, column=6).value
    Press_time_proportion = button_control_sheet.cell(row=5, column=7).value or button_control_sheet.cell(row=5, column=6).value
    Total_time = button_control_sheet.cell(row=6, column=7).value or button_control_sheet.cell(row=6, column=6).value
    Total_time = int(Total_time / 100)
    Press_numbers = button_control_sheet.cell(row=7, column=7).value or button_control_sheet.cell(row=7, column=6).value

    payload_list = int_number_to_int_list(Button_number,2) + int_number_to_int_list(Press_time_proportion,1) + int_number_to_int_list(Total_time,2) + int_number_to_int_list(Press_numbers,2)

    button_control_request_command_list = int_number_to_int_list(len(payload_list),2) + payload_list

    button_control_request_command_list.insert(0, int(moudle_id))

    calculate_crc_add_magicNumber(button_control_request_command_list)

    workbook.close()

    return {'Button Control':button_control_request_command_list}

def connect_and_set_mode():
    # 定义返回值
    success = "Success"
    connection_error = "Connection Error"
    mode_error = "Failed to set ITF mode"

    SERVER_ADDRESS = '192.168.3.11'
    SERVER_PORT = 6666

    hmi_exit_string = "5A 06 01 00 00 F8"
    hmi_exit_command = bytes.fromhex(hmi_exit_string.replace(" ", ""))

    # 创建socket对象
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    try:
        # 尝试连接服务器
        sock.connect((SERVER_ADDRESS, SERVER_PORT))

        # 发送数据
        sock.sendall(hmi_exit_command)

        # 准备接收数据
        data_recieve_hex = sock.recv(8)     #ACK

        data_recieve_string = ' '.join(['{:02X}'.format(x) for x in data_recieve_hex])

        print(data_recieve_string)

        if data_recieve_hex[5] == 0x25:
            return success

        if data_recieve_hex[5] != 0x01:
            return mode_error

        data_recieve_hex2 = sock.recv(8)    #Notify

        data_recieve_string2 = ' '.join(['{:02X}'.format(x) for x in data_recieve_hex2])

        print(data_recieve_string2)

        if data_recieve_hex2[5] != 0x01:
            return mode_error

        return success

    except socket.error as serr:
        # 如果连接失败
        return connection_error
    finally:
        # 关闭socket
        sock.close()

#print(connect_and_set_mode())

def send_control_command():
    SERVER_ADDRESS = '192.168.3.11'
    SERVER_PORT = 6666

    command_dict = generate_button_control_request_command("Lite DVF Configuration File_v0.5.xlsx", "00")

    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    try:
        # 尝试连接服务器
        sock.connect((SERVER_ADDRESS, SERVER_PORT))

        #print([format(num, '02X') for num in command_dict['Button Control']])
        # 发送数据
        sock.sendall(bytes(command_dict['Button Control']))

        # 准备接收数据
        data_recieve_hex = sock.recv(10)     #ACK

        data_recieve_string = ' '.join(['{:02X}'.format(x) for x in data_recieve_hex])

        print(data_recieve_string)

        if data_recieve_hex[5] != 0x01:
            return "ACK Error"

        data_recieve_hex2 = sock.recv(10)    #Notify

        data_recieve_string2 = ' '.join(['{:02X}'.format(x) for x in data_recieve_hex2])

        print(data_recieve_string2)

        if data_recieve_hex2[5] != 0x01:
            return "Notify Error"

        return "Success"

    except socket.error as serr:
        # 如果连接失败
        return "Connection Error"
    finally:
        # 关闭socket
        sock.close()

#print(connect_and_set_mode())

print(send_control_command())