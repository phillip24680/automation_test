from openpyxl import load_workbook

#V1.0

# 加载Excel文件
workbook = load_workbook(filename='Lite DVF Configuration File_v0.2.xlsx')

# 选择工作表
global_sheet = workbook['Global Configuration']

test_suite_fast = []  #存储执行fast configure的test suite
# 初始化行号
row_num = 4  # 从E4开始，行号为4
# 循环读取，直到遇到空值
while True:
    # 获取单元格值
    cell_value = global_sheet.cell(row=row_num, column=5).value  # 列号5对应E列
    if cell_value is None:  # 如果单元格为空，则停止循环
        break
    elif cell_value == 'Y':
        test_suite = global_sheet.cell(row=row_num, column=1).value
        test_suite_fast.append(test_suite)
        print(test_suite_fast[row_num - 4])
    else:
        print(global_sheet.cell(row=row_num, column=5).value)
    # 行号递增，移动到下一行
    row_num += 1


