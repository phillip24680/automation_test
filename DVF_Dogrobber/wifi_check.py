from openpyxl import load_workbook
import struct

def string_to_int_list(string):
    return [ord(char) for char in string]

def int_number_to_int_list(string_number, byte_length):
    """
    Convert an integer to a hexadecimal string in little-endian format with specified byte length.
    """
    packed_bytes = struct.pack(f'<{byte_length}s', struct.pack('<i', string_number)[:byte_length])
    return [byte for byte in packed_bytes]

def set_bits_from_reversed_string(bit_positions_str):
    # 将逗号分隔的字符串转换为整数列表，并确保在有效范围内
    values = [16 - int(position.strip()) for position in bit_positions_str.split(',') if
              1 <= int(position.strip()) <= 14]
    # 确保位置是唯一的并排序
    values = sorted(set(values))

    # 初始化一个16位全为0的二进制数
    bit_string = '0' * 16

    # 遍历有效值（逆序处理values，设置的是低位到高位）
    for value in values:
        bit_string = bit_string[:value] + '1' + bit_string[value + 1:]

    # 分割二进制字符串，前8位为高位B，后8位为低位A
    high_8_bits = bit_string[:8]
    low_8_bits = bit_string[8:]

    return [int(low_8_bits, 2), int(high_8_bits, 2)]

def calculate_crc_add_magicNumber(int_data_list):
    # 将十六进制数求和
    sum_value = sum(int_data_list)
    #sum_value = hex(sum_value)
    # 按位取反操作
    crc_result_0 = sum_value ^ 0xFF
    # 取后8位
    crc_result = crc_result_0 & 0xFF
    int_data_list.append(crc_result)
    int_data_list.insert(0, 90)

    return int_data_list

def generate_wifi_scan_request_command(wifi_check_sheet, moudle_id):

    Wifi_scan_subcommand = wifi_check_sheet.cell(row=4, column=2).value
    Timeout = wifi_check_sheet.cell(row=4, column=7).value
    Element_number = wifi_check_sheet.cell(row=5, column=7).value
    SSID_length = len(wifi_check_sheet.cell(row=7, column=7).value)
    SSID_string = wifi_check_sheet.cell(row=7, column=7).value
    Channel_list = wifi_check_sheet.cell(row=8, column=7).value

    #print(Channel_list)

    payload_list = [int(Wifi_scan_subcommand)]+int_number_to_int_list(Timeout, 2)+int_number_to_int_list(Element_number,1)+int_number_to_int_list(SSID_length,1)+string_to_int_list(SSID_string)+set_bits_from_reversed_string(Channel_list)

    wifi_scan_request_command_list = int_number_to_int_list(len(payload_list),2) + payload_list

    wifi_scan_request_command_list.insert(0,int(moudle_id))

    calculate_crc_add_magicNumber(wifi_scan_request_command_list)

    #print(wifi_scan_request_command_list)

    #print([format(num, '02X') for num in wifi_scan_request_command_list])

    return wifi_scan_request_command_list


def generate_wifi_check_request_command(excel_file, moudle_id):
    wifi_check_command_dict = {}
    # 加载Excel文件
    workbook = load_workbook(filename=excel_file, data_only=True)
    # 选择工作表
    wifi_check_sheet = workbook['Wi-Fi Check']

    global_fast_configuration = wifi_check_sheet.cell(row=2, column=7).value
    #global_execution_option = wifi_check_sheet.cell(row=2, column=6).value
    #print(global_fast_configuration)
    if global_fast_configuration == "Y":
        wifi_scan_fast_config = wifi_check_sheet.cell(row=4, column=4).value
        #以下为遍历所有子测试项的快速配置
        if wifi_scan_fast_config == "Y":
            wifi_check_command_dict["Wi-Fi Scanning"] = generate_wifi_scan_request_command(wifi_check_sheet, moudle_id)

    if global_fast_configuration == "N":
        #wifi_check_command_dict["test scope"] = []
        wifi_scan_execution_option = wifi_check_sheet.cell(row=4, column=3).value
        if wifi_scan_execution_option == "Y":
            wifi_check_command_dict["test scope"] = [int(moudle_id), int(wifi_check_sheet.cell(row=4, column=2).value)]
            wifi_check_command_dict["Wi-Fi Scanning"] = generate_wifi_scan_request_command(wifi_check_sheet, moudle_id)

    workbook.close()
    return wifi_check_command_dict

#print(generate_wifi_check_request_command("Lite DVF Configuration File_v0.2.xlsx", "03"))