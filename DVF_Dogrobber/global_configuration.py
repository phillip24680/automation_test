from openpyxl import load_workbook
import struct

from wifi_check import generate_wifi_check_request_command

#V1.0

def calculate_crc_add_magicNumber(hex_data_list):
    # 将十六进制数求和
    sum_value = sum(hex_data_list)
    #sum_value = hex(sum_value)
    # 按位取反操作
    crc_result_0 = sum_value ^ 0xFF
    # 取后8位
    crc_result = crc_result_0 & 0xFF
    hex_data_list.append(crc_result)
    hex_data_list.insert(0, int('90'))

    return hex_data_list

def get_test_suite_imformation(excel_file):
    # 加载Excel文件
    workbook = load_workbook(filename=excel_file)
    # 选择工作表
    global_sheet = workbook['Global Configuration']
    data_dict = {}
    for row_num in range(1,30):
        test_suite_title = global_sheet.cell(row=row_num, column=1).value
        if test_suite_title == "Test Suite Name":
            for row_sort in range(1,30):
                test_suite_name = global_sheet.cell(row=row_num + row_sort, column=1).value
                if test_suite_name is not None:
                    #test_suite_dict = {}
                    for column_num in range(1, 10):
                        command_ID_title = global_sheet.cell(row=row_num, column=column_num).value
                        if command_ID_title == "Corresponding Command ID":
                            test_suite_command_ID = global_sheet.cell(row=row_num + row_sort, column=column_num).value
                            break
                    for column_num in range(1, 10):
                        execution_option_title = global_sheet.cell(row=row_num, column=column_num).value
                        if execution_option_title == "Global Execution Option":
                            test_suite_execution_option = global_sheet.cell(row=row_num + row_sort, column=column_num).value
                            break
                    for column_num in range(1, 10):
                        fast_config_title = global_sheet.cell(row=row_num, column=column_num).value
                        if fast_config_title == "Global Fast Config":
                            test_suite_fast_config = global_sheet.cell(row=row_num + row_sort, column=column_num).value
                            break
                    test_suite_dict = {"Corresponding Command ID":test_suite_command_ID,"Global Execution Option":test_suite_execution_option,"Global Fast Config":test_suite_fast_config}
                    data_dict[test_suite_name] = test_suite_dict
                else:
                    return data_dict
            return data_dict

def check_whether_fast_config_exists(data):
    for key, value in data.items():
        if 'Global Fast Config' in value and value['Global Fast Config'] == 'Y':
            return True
    return False

def int_number_to_int_list(string_number, byte_length):
    """
    将整数转换为具有指定字节长度的小端格式的十六进制字符串。
    """
    packed_bytes = struct.pack(f'<{byte_length}s', struct.pack('<i', string_number)[:byte_length])
    return [byte for byte in packed_bytes]

def process_test_scope_commands(test_scope_list):
    test_scope_list.insert(0, 1)  #插入整数1，代表test scope的config subcommand
    payload_length_list = int_number_to_int_list(len(test_scope_list), 2)
    test_scope_list[:0] = payload_length_list  # 在列表开始处插入payload_length_list的元素
    test_scope_list.insert(0, 6)  #插入整数6，代表command为06，即config
    calculate_crc_add_magicNumber(test_scope_list)

def generate_commands_dict(excel_file):
    test_suite_imformation = get_test_suite_imformation(excel_file)
    if check_whether_fast_config_exists(test_suite_imformation):
        fast_commands_dict = {}
        #进入fast config选项
        if test_suite_imformation["Wi-Fi Check"]["Global Fast Config"] == "Y":
            wifi_fast_commands_dict = generate_wifi_check_request_command(excel_file, test_suite_imformation["Wi-Fi Check"]["Corresponding Command ID"])
            fast_commands_dict.update(wifi_fast_commands_dict)
        #print([format(num, '02X') for num in fast_commands_dict["Wi-Fi Scanning"]])
        return fast_commands_dict
    else:
        normal_commands_dict = {"test scope":[]}
        #进入normal config选项
        if test_suite_imformation["Wi-Fi Check"]["Global Execution Option"] == "Y":
            wifi_normal_commands_dict = generate_wifi_check_request_command(excel_file, test_suite_imformation["Wi-Fi Check"]["Corresponding Command ID"])
            normal_commands_dict["test scope"] = normal_commands_dict["test scope"] + wifi_normal_commands_dict["test scope"]
            wifi_normal_commands_dict.pop("test scope")  #删除字典中"test scope"元素
            normal_commands_dict.update(wifi_normal_commands_dict)

        process_test_scope_commands(normal_commands_dict["test scope"])
        #print([format(num, '02X') for num in normal_commands_dict["test scope"]])
        #print([format(num, '02X') for num in normal_commands_dict["Wi-Fi Scanning"]])
        return normal_commands_dict

#print(generate_commands_dict("Lite DVF Configuration File_v0.2.xlsx"))




